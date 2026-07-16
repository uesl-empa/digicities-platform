# SPDX-License-Identifier: Apache-2.0
# Copyright © 2026, Empa, James Allan, Reto Fricker

"""Generate the Alpine Village Excel template for the Replica Builder importer.

Run from the repo root:

    python tutorial/sample_data/_build_replica_template.py

Writes ``tutorial/sample_data/alpine_village_replica_template.xlsx`` —
the file consumed by the Streamlit "Replica Builder → Excel Import" tab and
walked through in ``tutorial/09_excel_import.ipynb``.

The .xlsx is committed alongside this script. Re-run only when the
template content changes; reviewers can read the script to see the diff
since the binary .xlsx isn't reviewable on its own.

Header convention (consumed by ``backend.replica_builder.utils.create_class_and_attribute_graph.process_excel_to_ttl``):

    Row 1: attribute name (column "id" marks the instance ID column)
    Row 2: attribute type (Physical, UnitBasedCost, ClassObject, ...)
    Row 3: unit            (QUDT short code, e.g. KiloW, M2)
    Row 4: unit_y          (Curve y-axis unit, or denominator for CustomPhysicalRatio)
    Row 5: currency        (ISO code for cost attributes)
    Row 6: predicate       (dici_onto property name for ClassObject attributes)
    Row 7+: instance rows
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_PATH = Path(__file__).resolve().parent / "alpine_village_replica_template.xlsx"

HEADER_FILL = PatternFill(start_color="E8EEF7", end_color="E8EEF7", fill_type="solid")
TYPE_FILL = PatternFill(start_color="F4F4F4", end_color="F4F4F4", fill_type="solid")
HEADER_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def write_sheet(ws, columns, rows):
    """columns: list of {name, type?, unit?, unit_y?, currency?, predicate?}.
    rows: list of dicts keyed by column name.
    """
    for col_idx, col in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col.get("name", ""))
        ws.cell(row=2, column=col_idx, value=col.get("type", ""))
        ws.cell(row=3, column=col_idx, value=col.get("unit", ""))
        ws.cell(row=4, column=col_idx, value=col.get("unit_y", ""))
        ws.cell(row=5, column=col_idx, value=col.get("currency", ""))
        ws.cell(row=6, column=col_idx, value=col.get("predicate", ""))
        for r in (1, 2):
            cell = ws.cell(row=r, column=col_idx)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        for r in (3, 4, 5, 6):
            ws.cell(row=r, column=col_idx).fill = TYPE_FILL

    for row_idx, row in enumerate(rows, start=7):
        for col_idx, col in enumerate(columns, start=1):
            val = row.get(col["name"])
            if val is not None:
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = WRAP

    for col_idx, col in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        widest = len(str(col.get("name", "")))
        for row in rows:
            v = row.get(col["name"])
            if v is not None:
                widest = max(widest, min(len(str(v)), 40))
        ws.column_dimensions[letter].width = max(widest + 2, 12)

    ws.freeze_panes = "B7"


def build():
    wb = Workbook()
    wb.remove(wb.active)

    # ------------------------------------------------------------------
    # EnergyConsumer — three buildings demonstrating Physical, Categorical,
    # Event, ClassObject, Historic time-series, Identifier, and Annotation,
    # plus per-attribute datasource columns linking to the Reference sheet.
    # ------------------------------------------------------------------
    ws = wb.create_sheet("EnergyConsumer")
    write_sheet(
        ws,
        columns=[
            {"name": "id"},
            {"name": "floorArea", "type": "Physical", "unit": "M2"},
            {"name": "floorArea_datasource"},
            {"name": "electricityDemand", "type": "Physical", "unit": "KiloW-HR"},
            {"name": "electricityDemand_datasource"},
            {"name": "buildingType", "type": "Categorical"},
            {"name": "yearBuilt", "type": "Event"},
            {"name": "locatedIn", "type": "ClassObject", "predicate": "locatedIn"},
            {"name": "monthlyDemandProfile", "type": "Historic", "unit": "KiloW-HR"},
            {"name": "buildingId", "type": "Identifier"},
            {"name": "label", "type": "Annotation"},
        ],
        rows=[
            {
                "id": "BuildingA",
                "floorArea": 180.0,
                "floorArea_datasource": "swiss_energy_atlas_2024",
                "electricityDemand": 4800.0,
                "electricityDemand_datasource": "swiss_energy_atlas_2024",
                "buildingType": "SingleFamilyHouse",
                "yearBuilt": 1970,
                "locatedIn": "Location/AlpineValley",
                "monthlyDemandProfile": "resources/buildingA_monthly_demand.csv",
                "buildingId": "BLDG-A-001",
                "label": "Single-family house with rooftop PV + battery",
            },
            {
                "id": "BuildingB",
                "floorArea": 850.0,
                "electricityDemand": 22000.0,
                "buildingType": "ApartmentBlock",
                "yearBuilt": 1985,
                "locatedIn": "Location/AlpineValley",
                "monthlyDemandProfile": "resources/buildingB_monthly_demand.csv",
                "buildingId": "BLDG-B-001",
                "label": "Apartment block heated by an air-source heat pump",
            },
            {
                "id": "BuildingC",
                "floorArea": 165.0,
                "electricityDemand": 3500.0,
                "buildingType": "SingleFamilyHouse",
                "yearBuilt": 1992,
                "locatedIn": "Location/AlpineValley",
                "buildingId": "BLDG-C-001",
                "label": "Single-family house, passive consumer",
            },
        ],
    )

    # ------------------------------------------------------------------
    # EnergyConverter — heat pump and gas boiler demonstrate UnitBasedCost,
    # SimpleCost, dimensionless Physical, Curve, and ClassObject linking
    # back to the EnergyConsumer sheet.
    # ------------------------------------------------------------------
    ws = wb.create_sheet("EnergyConverter")
    write_sheet(
        ws,
        columns=[
            {"name": "id"},
            {"name": "nominalPower", "type": "Physical", "unit": "KiloW"},
            {"name": "efficiency", "type": "Physical"},
            {"name": "capEx", "type": "UnitBasedCost", "unit": "KiloW", "currency": "CHF"},
            {"name": "annualMaintenance", "type": "SimpleCost", "currency": "CHF"},
            {"name": "partLoadEfficiency", "type": "Curve", "unit": "PERCENT", "unit_y": "PERCENT"},
            {"name": "installedAt", "type": "ClassObject", "predicate": "installedAt"},
            {"name": "label", "type": "Annotation"},
        ],
        rows=[
            {
                "id": "HeatPump_B",
                "nominalPower": 12.0,
                "efficiency": 3.8,
                "capEx": 1800.0,
                "annualMaintenance": 250.0,
                "partLoadEfficiency": "[(25.0,280.0);(50.0,360.0);(75.0,410.0);(100.0,390.0)]",
                "installedAt": "EnergyConsumer/BuildingB",
                "label": "Air-source heat pump for Building B",
            },
            {
                "id": "Boiler_C",
                "nominalPower": 15.0,
                "efficiency": 0.94,
                "capEx": 350.0,
                "annualMaintenance": 180.0,
                "installedAt": "EnergyConsumer/BuildingC",
                "label": "Retrofit gas boiler at Building C",
            },
        ],
    )

    # ------------------------------------------------------------------
    # Network — the regional electricity grid. Demonstrates
    # CustomPhysicalRatio for prices that are ratios of two QUDT units
    # (no single QUDT IRI fits, so the importer emits dici_onto:hasUnitLabel
    # with the composite "CHF/KiloW-HR" string).
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Network")
    write_sheet(
        ws,
        columns=[
            {"name": "id"},
            {"name": "exportLimit", "type": "Physical", "unit": "KiloW"},
            {"name": "importPrice", "type": "CustomPhysicalRatio", "unit": "CHF", "unit_y": "KiloW-HR"},
            {"name": "label", "type": "Annotation"},
        ],
        rows=[
            {
                "id": "Grid",
                "exportLimit": 50.0,
                "importPrice": 0.25,
                "label": "Regional electricity grid connection",
            },
        ],
    )

    # ------------------------------------------------------------------
    # Location — the alpine valley referenced by every Building's
    # locatedIn ClassObject. One sheet per dici_onto class.
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Location")
    write_sheet(
        ws,
        columns=[
            {"name": "id"},
            {"name": "label", "type": "Annotation"},
        ],
        rows=[
            {"id": "AlpineValley", "label": "Alpine Valley"},
        ],
    )

    # ------------------------------------------------------------------
    # Reference — citation sheet. IDs here can appear in any
    # *_datasource column; the importer rewrites those into
    # prov:wasDerivedFrom links to a typed dici_onto:Reference instance.
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Reference")
    write_sheet(
        ws,
        columns=[
            {"name": "id"},
            {"name": "description"},
            {"name": "ReferenceType"},
            {"name": "URL"},
            {"name": "AccessDate"},
            {"name": "comment"},
        ],
        rows=[
            {
                "id": "swiss_energy_atlas_2024",
                "description": "Swiss Federal Energy Atlas 2024",
                "ReferenceType": "Dataset",
                "URL": "https://www.uvek-gis.admin.ch/BFE/storymaps/EE_Energieatlas/",
                "AccessDate": "2024-03-15",
                "comment": "Federal building-stock electricity demand database",
            },
        ],
    )

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"wrote {OUT_PATH}")


if __name__ == "__main__":
    build()
