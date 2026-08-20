"""Replica Builder endpoints — the reusable path: a digital-replica workbook
(.xlsx) to instance TTL via the platform's own converter, plus reading back the
workspace's current replica TTL for Preview & Export.

The in-app Instances/Attributes/Links editor (session-coupled TTL generation)
is a later chunk; this covers Excel Import + Preview & Export + config.
"""
from __future__ import annotations

import tempfile
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel

from backend.workspace import WorkspaceContext

from .deps import get_ctx, ws_root

router = APIRouter(prefix="/api/workspaces/{workspace_id}/replica", tags=["replica"])

_PROJECT_PREFIX = "https://digicities.info/proj"



def _project_uri(ctx: WorkspaceContext) -> str:
    return f"{_PROJECT_PREFIX}/{ctx.id}"


@router.get("/config")
def config(ctx: WorkspaceContext = Depends(get_ctx)) -> dict[str, str]:
    return {"workspace": ctx.id, "project_uri": _project_uri(ctx)}


@router.get("/ttl")
def replica_ttl(ctx: WorkspaceContext = Depends(get_ctx)) -> dict[str, Any]:
    """The workspace's current replica TTL (Preview & Export)."""
    out = ws_root(ctx) / "ingestion" / "output"
    files = sorted(out.glob("*.ttl")) if out.exists() else []
    if not files:
        return {"ttl": "", "file": None}
    f = files[0]
    return {"ttl": f.read_text(encoding="utf-8"), "file": f.name}


@router.post("/import")
async def import_workbook(
    file: UploadFile = File(...),
    ctx: WorkspaceContext = Depends(get_ctx),
) -> dict[str, Any]:
    """Convert an uploaded digital-replica workbook (.xlsx) to instance TTL via
    ``process_excel_to_ttl`` and persist it to the workspace's ingestion output."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Upload a .xlsx digital-replica workbook.")

    root = ws_root(ctx)
    in_dir = root / "ingestion" / "input"
    out_dir = root / "ingestion" / "output"
    in_dir.mkdir(parents=True, exist_ok=True)
    out_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path = in_dir / file.filename
    xlsx_path.write_bytes(await file.read())
    ttl_path = out_dir / f"{ctx.id}.ttl"

    from backend.replica_builder.utils.create_class_and_attribute_graph import process_excel_to_ttl

    try:
        process_excel_to_ttl(_project_uri(ctx), str(xlsx_path), str(ttl_path), uri_mode="default")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Conversion failed: {exc}") from exc

    ttl = ttl_path.read_text(encoding="utf-8") if ttl_path.exists() else ""
    return {"file": ttl_path.name, "triples_preview": ttl[:4000], "ttl": ttl, "chars": len(ttl)}


# ── in-app builder: a replica model (classes + typed attribute columns + instance
# rows) -> a workbook -> process_excel_to_ttl. Same 6-row header the agent writes. ──
class Column(BaseModel):
    name: str
    type: str | None = None
    unit: str | None = None
    unit_y: str | None = None
    currency: str | None = None
    predicate: str | None = None


class Component(BaseModel):
    cls: str
    columns: list[Column] = []
    rows: list[dict[str, Any]] = []


class ReplicaSpec(BaseModel):
    components: list[Component]
    persist: bool = True


def _build_workbook(spec: ReplicaSpec, path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    for comp in spec.components:
        ws = wb.create_sheet(title=comp.cls[:31])
        columns = [Column(name="id")] + comp.columns
        for c, col in enumerate(columns, start=1):
            ws.cell(row=1, column=c, value=col.name)
            ws.cell(row=2, column=c, value=col.type)
            ws.cell(row=3, column=c, value=col.unit)
            ws.cell(row=4, column=c, value=col.unit_y)
            ws.cell(row=5, column=c, value=col.currency)
            ws.cell(row=6, column=c, value=col.predicate)
        for r, row in enumerate(comp.rows, start=7):
            for c, col in enumerate(columns, start=1):
                ws.cell(row=r, column=c, value=row.get(col.name))
    wb.save(str(path))


@router.post("/generate")
def generate(spec: ReplicaSpec, ctx: WorkspaceContext = Depends(get_ctx)) -> dict[str, Any]:
    """Build a workbook from the in-app replica model and convert it to instance TTL."""
    if not spec.components:
        raise HTTPException(status_code=400, detail="Add at least one component class.")
    root = ws_root(ctx)
    (root / "ingestion" / "input").mkdir(parents=True, exist_ok=True)
    (root / "ingestion" / "output").mkdir(parents=True, exist_ok=True)
    xlsx = (root / "ingestion" / "input" / f"{ctx.id}.xlsx") if spec.persist \
        else Path(tempfile.mkdtemp()) / "replica.xlsx"
    ttl_path = (root / "ingestion" / "output" / f"{ctx.id}.ttl") if spec.persist \
        else Path(tempfile.mkdtemp()) / "replica.ttl"
    _build_workbook(spec, xlsx)

    from backend.replica_builder.utils.create_class_and_attribute_graph import process_excel_to_ttl

    try:
        process_excel_to_ttl(_project_uri(ctx), str(xlsx), str(ttl_path), uri_mode="default")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Generation failed: {exc}") from exc
    ttl = ttl_path.read_text(encoding="utf-8") if ttl_path.exists() else ""
    return {"ttl": ttl, "chars": len(ttl), "persisted": spec.persist}
