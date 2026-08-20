# SPDX-License-Identifier: Apache-2.0
# Copyright © 2026, Empa, James Allan, Reto Fricker

"""ReplicaDraft — the serializable input contract of the replica generator.

This formalizes the schema ``apps/api/replica.py`` already accepts on
``POST /replica/generate`` (adopted as-is from the colleague's ``ReplicaSpec``):
a list of components, each a class name plus typed attribute *columns*
(``{name, type, unit, unit_y, currency, predicate}`` — mirroring the 6-row
workbook header) and plain instance *rows* (dicts keyed by column name).

The one extension over the wire schema is the optional per-column ``key``: the
row-dict key the cell value is read from, defaulting to ``name``. The workbook
format identifies time-series variants by giving several columns the SAME name
with different type rows (``X``/Historic, ``X``/Future …), which a plain dict
row cannot express — ``key`` disambiguates them (used by
:meth:`ReplicaDraft.from_session_state`). API requests never need to set it.

The draft's one consumer is :func:`build_workbook` (moved here from
``apps/api/replica.py::_build_workbook``): draft → 6-row-header workbook →
``process_excel_to_ttl``. That keeps the Excel converter the single parser of
replica models.

Constructors:

* :meth:`from_request` — API request dicts (``spec.model_dump()``).
* :meth:`from_session_state` / :meth:`from_instances` — the Streamlit session
  model (``replica_instances``). Links (``replica_links``) are the
  system_description graph and have no workbook representation, so they are
  deliberately not part of the draft.
* :meth:`to_dict` / :meth:`from_dict` — plain-dict round trip.
"""
from __future__ import annotations

from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence


@dataclass
class ReplicaColumn:
    name: str
    type: Optional[str] = None        # header row 2: attribute type
    unit: Optional[str] = None        # header row 3: QUDT unit (x-unit for Curve)
    unit_y: Optional[str] = None      # header row 4: y-unit (Curve) / denominator
    currency: Optional[str] = None    # header row 5: currency (cost types)
    predicate: Optional[str] = None   # header row 6: predicate (ClassObject)
    key: Optional[str] = None         # row-dict lookup key; defaults to name


@dataclass
class ReplicaComponent:
    cls: str
    columns: List[ReplicaColumn] = field(default_factory=list)
    rows: List[Dict[str, Any]] = field(default_factory=list)


@dataclass
class ReplicaDraft:
    components: List[ReplicaComponent] = field(default_factory=list)

    # ── serialization ─────────────────────────────────────────────────────
    def to_dict(self) -> dict:
        return asdict(self)

    @classmethod
    def from_dict(cls, data: Mapping[str, Any]) -> "ReplicaDraft":
        return cls.from_request(dict(data).get("components", []))

    # ── constructors ──────────────────────────────────────────────────────
    @classmethod
    def from_request(cls, components: Sequence[Mapping[str, Any]]) -> "ReplicaDraft":
        """Build a draft from API-request shapes (``ReplicaSpec.components``).

        Raises ``ValueError`` when a component has no ``cls`` or a column has
        no ``name`` — the workbook sheet/header cannot be written without them.
        """
        norm: List[ReplicaComponent] = []
        for comp in components:
            comp = dict(comp)
            cls_name = comp.get("cls")
            if not cls_name:
                raise ValueError("every component needs a 'cls' (the component class name)")
            columns: List[ReplicaColumn] = []
            for col in comp.get("columns", []) or []:
                col = dict(col)
                if not col.get("name"):
                    raise ValueError(f"component '{cls_name}': every column needs a 'name'")
                columns.append(ReplicaColumn(
                    name=col["name"],
                    type=col.get("type"),
                    unit=col.get("unit"),
                    unit_y=col.get("unit_y"),
                    currency=col.get("currency"),
                    predicate=col.get("predicate"),
                    key=col.get("key"),
                ))
            norm.append(ReplicaComponent(
                cls=cls_name,
                columns=columns,
                rows=[dict(r) for r in comp.get("rows", []) or []],
            ))
        return cls(components=norm)

    @classmethod
    def from_session_state(cls, state: Mapping[str, Any]) -> "ReplicaDraft":
        """Build a draft from Streamlit session state (or any mapping with the
        same keys): ``replica_instances`` grouped by component type, with the
        project URI (``replica_project_uri``) used to relativize ClassObject
        targets back to the ``Sheet/id`` cell convention."""
        return cls.from_instances(
            list(state.get("replica_instances", [])),
            project_uri=state.get("replica_project_uri"),
        )

    @classmethod
    def from_instances(cls, instances: Sequence[Any],
                       project_uri: Optional[str] = None) -> "ReplicaDraft":
        """Build a draft from ComponentInstance objects (session model or the
        graph loader's parse-back). Attribute dicts map back onto workbook
        columns exactly inversely to how the converter reads them."""
        by_type: Dict[str, List[Any]] = {}
        for inst in instances:
            by_type.setdefault(inst.component_type, []).append(inst)

        components: List[ReplicaComponent] = []
        for comp_type in by_type:
            columns: Dict[str, ReplicaColumn] = {}   # keyed by row-lookup key
            rows: List[Dict[str, Any]] = []

            def _column(key: str, **spec: Any) -> ReplicaColumn:
                if key not in columns:
                    name = spec.pop("name")
                    columns[key] = ReplicaColumn(
                        name=name, key=key if key != name else None, **spec)
                return columns[key]

            for inst in by_type[comp_type]:
                row: Dict[str, Any] = {"id": inst.id}

                for key, value in (inst.annotations or {}).items():
                    if key == "label":
                        continue  # already the instance label / id
                    _column(key, name=key, type="Annotation")
                    row[key] = value

                for predicate, target in (getattr(inst, "class_objects", None) or {}).items():
                    _column(predicate, name=predicate, type="ClassObject",
                            predicate=predicate)
                    cell = target
                    if project_uri and isinstance(target, str) \
                            and target.startswith(f"{project_uri}/"):
                        cell = target[len(project_uri) + 1:]
                    row[predicate] = cell

                for name, data in (inst.attributes or {}).items():
                    cells = _attribute_cells(name, data)
                    for key, (spec, value) in cells.items():
                        _column(key, **spec)
                        row[key] = value

                rows.append(row)

            components.append(ReplicaComponent(
                cls=comp_type, columns=list(columns.values()), rows=rows))

        return cls(components=components)


def _attribute_cells(name: str, data: Mapping[str, Any]) -> Dict[str, Any]:
    """Map one session attribute dict to workbook cells: {row_key: (column
    spec kwargs, cell value)}. Inverse of the converter's column handling."""
    attr_type = data.get("type", "Physical")
    cells: Dict[str, Any] = {}

    def put(key: str, value: Any, **spec: Any) -> None:
        if value is None or value == "":
            return
        spec.setdefault("name", name)
        cells[key] = (spec, value)

    if attr_type in ("Physical", "Dynamic", "Geospatial"):
        put(name, data.get("value"), type="Physical" if attr_type == "Dynamic" else attr_type,
            unit=data.get("unit"))
        # Time-series variants: same column name, distinct type row.
        put(f"{name}::historic", data.get("historic_reference"),
            type="Historic", unit=data.get("unit"))
        put(f"{name}::future", data.get("future_reference"),
            type="Future", unit=data.get("unit"))
        put(f"{name}::live", data.get("live_reference"),
            type="Live", unit=data.get("unit"))
        # Legacy Excel-importer shape: one variant per attribute dict.
        if data.get("time_series_type") and data.get("reference"):
            put(f"{name}::{str(data['time_series_type']).lower()}",
                data.get("reference"), type=data["time_series_type"],
                unit=data.get("unit"))
    elif attr_type == "Categorical":
        put(name, data.get("category_value"), type="Categorical")
    elif attr_type == "Event":
        put(name, data.get("temporal_value"), type="Event")
    elif attr_type == "SimpleCost":
        put(name, data.get("value"), type="SimpleCost", currency=data.get("currency"))
    elif attr_type == "UnitBasedCost":
        put(name, data.get("value"), type="UnitBasedCost",
            unit=data.get("unit"), currency=data.get("currency"))
    elif attr_type == "Curve":
        put(name, data.get("data_points"), type="Curve",
            unit=data.get("x_unit"), unit_y=data.get("y_unit"))
    elif attr_type == "Resource":
        put(name, data.get("data_path"), type="Resource")
    elif attr_type == "SimpleValue":
        put(name, data.get("value"), type="SimpleValue")
    elif attr_type == "CustomPhysicalRatio":
        custom = str(data.get("custom_unit", "") or "")
        num, _, den = custom.partition("/")
        put(name, data.get("value"), type="CustomPhysicalRatio",
            unit=num or None, unit_y=den or None)
    elif attr_type == "Identifier":
        put(name, data.get("identifier_value"), type="Identifier")
    else:
        put(name, data.get("value"), type=attr_type, unit=data.get("unit"))

    if data.get("datasource"):
        put(f"{name}_datasource", data.get("datasource"), name=f"{name}_datasource")

    return cells


def build_workbook(draft: ReplicaDraft, path: Path | str) -> None:
    """Write the draft as a 6-row-header digital-replica workbook — the exact
    format ``process_excel_to_ttl`` (and the onboarding agent) reads.

    Moved from ``apps/api/replica.py::_build_workbook``; the only change is the
    per-column ``key`` row lookup (defaults to the column name).
    """
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    for comp in draft.components:
        ws = wb.create_sheet(title=comp.cls[:31])
        columns = [ReplicaColumn(name="id")] + list(comp.columns)
        for c, col in enumerate(columns, start=1):
            ws.cell(row=1, column=c, value=col.name)
            ws.cell(row=2, column=c, value=col.type)
            ws.cell(row=3, column=c, value=col.unit)
            ws.cell(row=4, column=c, value=col.unit_y)
            ws.cell(row=5, column=c, value=col.currency)
            ws.cell(row=6, column=c, value=col.predicate)
        for r, row in enumerate(comp.rows, start=7):
            for c, col in enumerate(columns, start=1):
                ws.cell(row=r, column=c, value=row.get(col.key or col.name))
    wb.save(str(path))


__all__ = ["ReplicaColumn", "ReplicaComponent", "ReplicaDraft", "build_workbook"]
